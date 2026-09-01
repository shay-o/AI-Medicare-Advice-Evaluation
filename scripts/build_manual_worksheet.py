"""Build the manual data-collection workbook for consumer chat product runs."""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = "/Users/jamesoreilly/Documents/Projects/AI-Medicare-Advice-Evaluator"
sys.path.insert(0, ROOT)
os.chdir(ROOT)
from src.grading_rubric import get_question_group_for_turn, get_study_question_number  # noqa: E402

OUT = "reference_material/manual_product_runs_worksheet.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
LOCK_FILL = PatternFill("solid", fgColor="EDEDED")      # pre-filled, do not edit
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")     # you fill this in
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="666666")
BODY = Font(name=FONT, size=10)
BLUE = Font(name=FONT, size=10, color="0000FF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SCORES = "accurate_complete,substantive_incomplete,not_substantive,incorrect"
YN = "Yes,No,Unclear"

wb = Workbook()

# ---------------------------------------------------------------- Instructions
ws = wb.active
ws.title = "Instructions"
ws.sheet_view.showGridLines = False
rows = [
    ("How to use this workbook", "h1"),
    ("", ""),
    ("This records manual runs through consumer chat products (ChatGPT.com, Claude.ai, Google.com) "
     "so they can be compared with the API results in this study.", "p"),
    ("", ""),
    ("Fill in one Session Log row per conversation, then the response rows for that scenario.", "p"),
    ("", ""),
    ("Colour key", "h2"),
    ("Grey cells are pre-filled. Do not edit them: the question wording is the study script.", "grey"),
    ("Yellow cells are for you to fill in.", "yellow"),
    ("", ""),
    ("Protocol: this is what makes the results comparable", "h2"),
    ("1. Start a NEW chat. Turn OFF memory, custom instructions and personalisation.", "p"),
    ("   A product that remembers earlier turns from a previous session is not answering the "
     "same question the study asked.", "sub"),
    ("2. Paste the opening statement (top of each scenario sheet) as the first message.", "p"),
    ("3. Ask each question EXACTLY as written, in order. Do not paraphrase or reword.", "p"),
    ("4. Do NOT add context, hints, or say you are testing it. Do not tell it to act as a "
     "Medicare counselor.", "p"),
    ("   Prompting it well would measure something other than what an ordinary person gets.", "sub"),
    ("5. If it asks where you live, give the location on the scenario sheet. If it asks anything "
     "else, answer only from the persona details, and give nothing that was not requested.", "p"),
    ("6. Paste the FULL response verbatim into response_text. Do not summarise or trim.", "p"),
    ("7. Record whether it searched the web, and paste any source links it showed.", "p"),
    ("", ""),
    ("Leave the score columns blank", "h2"),
    ("score and grader_notes are filled later by the grading pipeline, so that manual runs and "
     "API runs are judged by the same rubric. Do not score as you go.", "p"),
    ("", ""),
    ("Two turns are deliberately not scored", "h2"),
    ("They are marked 'not scored' in the question_group column: a location reply, and the second "
     "half of a two-part question, which the study scores inside its first part. Still ask them, "
     "and still record the response. They just do not get their own verdict.", "p"),
    ("", ""),
    ("One example row is provided on each scenario sheet, in blue, showing the expected format. "
     "Delete it before you start, or overwrite it.", "note"),
]
r = 1
for text, kind in rows:
    c = ws.cell(row=r, column=1, value=text)
    if kind == "h1":
        c.font = Font(name=FONT, bold=True, size=14, color="1F3864")
    elif kind == "h2":
        c.font = Font(name=FONT, bold=True, size=11, color="1F3864")
    elif kind == "sub":
        c.font = NOTE_FONT
    elif kind == "note":
        c.font = NOTE_FONT
    else:
        c.font = BODY
    if kind == "grey":
        c.fill = LOCK_FILL
    if kind == "yellow":
        c.fill = INPUT_FILL
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
ws.column_dimensions["A"].width = 110
for i in range(1, r):
    ws.row_dimensions[i].height = None

# ---------------------------------------------------------------- Session Log
sl = wb.create_sheet("Session Log")
sl_cols = [
    ("session_id", 14, "Your own id, e.g. CHATGPT-MO-01. Used in the scenario sheets."),
    ("date", 12, "Date of the conversation, YYYY-MM-DD."),
    ("product", 16, "ChatGPT.com, Claude.ai, or Google.com."),
    ("model_reported", 20, "What the product says it is running. Note this is often unreliable."),
    ("account_tier", 14, "Free, Plus, Pro, etc. Affects which model actually serves you."),
    ("scenario", 16, "medicare_only or dual_eligible."),
    ("memory_disabled", 16, "Yes/No. Must be Yes for the run to be comparable."),
    ("new_chat", 12, "Yes/No. Each scenario needs a fresh conversation."),
    ("logged_in", 12, "Yes/No."),
    ("location_given", 24, "The location you supplied when asked."),
    ("interruptions", 30, "Anything that broke protocol: a retry, a refusal, a lost connection."),
    ("notes", 40, "Anything else worth knowing about this session."),
]
for i, (h, w, note) in enumerate(sl_cols, 1):
    c = sl.cell(row=1, column=i, value=h)
    c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.comment = Comment(note, "worksheet")
    sl.column_dimensions[get_column_letter(i)].width = w
example = ["CHATGPT-MO-01", "2026-08-19", "ChatGPT.com", "GPT-5.6 Luna", "Plus", "medicare_only",
           "Yes", "Yes", "Yes", "Oakland, CA", "none", "example row, delete or overwrite"]
for i, v in enumerate(example, 1):
    c = sl.cell(row=2, column=i, value=v)
    c.font, c.border = BLUE, BORDER
for row in range(3, 60):
    for i in range(1, len(sl_cols) + 1):
        c = sl.cell(row=row, column=i)
        c.fill, c.border, c.font = INPUT_FILL, BORDER, BODY
sl.freeze_panes = "A2"
dv = DataValidation(type="list", formula1=f'"{YN.replace(",Unclear","")}"', allow_blank=True)
sl.add_data_validation(dv)
dv.add(f"G2:H{59}")
dvp = DataValidation(type="list", formula1='"ChatGPT.com,Claude.ai,Google.com"', allow_blank=True)
sl.add_data_validation(dvp)
dvp.add("C2:C59")
dvs = DataValidation(type="list", formula1='"medicare_only,dual_eligible"', allow_blank=True)
sl.add_data_validation(dvs)
dvs.add("F2:F59")

# ---------------------------------------------------------------- scenario sheets
COLS = [
    ("session_id", 15, True, "Must match a session_id in the Session Log."),
    ("turn", 6, False, "Conversation turn number. Pre-filled."),
    ("study_q", 8, False, "SHIP study question number from eAppendix 1 or 2. Blank means the study "
                          "does not number this turn."),
    ("question_group", 14, False, "Rubric group used for scoring. 'not scored' turns are still "
                                  "asked and recorded, they just get no verdict."),
    ("question_text", 62, False, "Ask this EXACTLY as written. Pre-filled from the study script."),
    ("response_text", 70, True, "Paste the full response verbatim. Do not summarise or trim."),
    ("searched_web", 13, True, "Yes/No/Unclear. Did the product search the web for this answer?"),
    ("citations", 45, True, "Source links it showed, one per line. Leave blank if none."),
    ("asked_for_location", 17, True, "Yes/No. Did it ask where you live on this turn?"),
    ("refused_or_deflected", 19, True, "Yes/No. Did it decline to answer or push you elsewhere?"),
    ("response_seconds", 15, True, "Roughly how long it took, if you can tell. Optional."),
    ("observations", 40, True, "Anything notable: hedging, a disclaimer, a follow-up question."),
    ("score", 22, True, "LEAVE BLANK. Filled later by the grading pipeline."),
    ("grader_notes", 30, True, "LEAVE BLANK. Filled later by the grading pipeline."),
]

SCEN = [("Medicare-Only", "medicare_only", "scenarios/medicare_only/all_questions.json"),
        ("Dual-Eligible", "dual_eligible", "scenarios/dual_eligible/all_questions.json")]

for title, key, path in SCEN:
    d = json.load(open(path))
    plan = d.get("plan_information") or {}
    doctor = (d.get("persona") or {}).get("primary_care_physician", "")
    persona = d.get("persona") or {}
    sh = wb.create_sheet(title)

    sh["A1"] = f"{title} scenario"
    sh["A1"].font = Font(name=FONT, bold=True, size=13, color="1F3864")
    sh["A2"] = "OPENING STATEMENT (send this first, before any question):"
    sh["A2"].font = Font(name=FONT, bold=True, size=10)
    sh["A3"] = d.get("opening_statement", "")
    sh["A3"].font = Font(name=FONT, size=10, italic=True)
    sh["A3"].fill = LOCK_FILL
    sh["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    sh.merge_cells("A3:E3")
    sh.row_dimensions[3].height = 30

    details = [f"Location to give if asked: {persona.get('location','')}",
               f"Plan asked about: {plan.get('plan_name','n/a')}",
               f"Doctor to ask about: {doctor or 'n/a'}",
               f"Persona: {str(persona.get('situation',''))[:150]}"]
    sh["A4"] = "  |  ".join(x for x in details if not x.endswith("n/a"))
    sh["A4"].font = NOTE_FONT
    sh["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    sh.merge_cells("A4:E4")
    sh.row_dimensions[4].height = 28

    hr = 6
    for i, (h, w, _inp, note) in enumerate(COLS, 1):
        c = sh.cell(row=hr, column=i, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.comment = Comment(note, "worksheet")
        sh.column_dimensions[get_column_letter(i)].width = w

    turns = d["scripted_turns"]
    r = hr + 1
    # one example row first
    ex = ["CHATGPT-MO-01", "", "", "", "", "(paste the model's full reply here)", "Yes",
          "https://www.medicare.gov/...", "No", "No", "12", "cited two sources", "", ""]
    for i, v in enumerate(ex, 1):
        c = sh.cell(row=r, column=i, value=v)
        c.font, c.border = BLUE, BORDER
        c.alignment = Alignment(wrap_text=True, vertical="top")
    sh.cell(row=r, column=5, value="(example row, delete or overwrite)").font = BLUE
    r += 1

    first_data = r
    for t_i, t in enumerate(turns, 1):
        msg = (t.get("user_message", "")
               .replace("[plan name]", plan.get("plan_name", ""))
               .replace("{plan_name}", plan.get("plan_name", ""))
               .replace("[doctor name]", doctor)
               .replace("{doctor_name}", doctor))
        g = get_question_group_for_turn(t_i, key)
        sq = get_study_question_number(t_i, key)
        vals = [None, t_i, sq if sq else "", g.group_id if g else "not scored", msg]
        for i, v in enumerate(vals, 1):
            c = sh.cell(row=r, column=i, value=v)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if i == 1:
                c.fill, c.font = INPUT_FILL, BODY
            else:
                c.fill = LOCK_FILL
                c.font = Font(name=FONT, size=10,
                              italic=(v == "not scored"),
                              color="999999" if v == "not scored" else "000000")
        for i in range(6, len(COLS) + 1):
            c = sh.cell(row=r, column=i)
            c.fill, c.border, c.font = INPUT_FILL, BORDER, BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
        sh.row_dimensions[r].height = 42
        r += 1

    last = r - 1
    sh.freeze_panes = "F7"
    dvs2 = DataValidation(type="list", formula1=f'"{SCORES}"', allow_blank=True)
    sh.add_data_validation(dvs2)
    dvs2.add(f"M{first_data}:M{last}")
    dvy = DataValidation(type="list", formula1=f'"{YN}"', allow_blank=True)
    sh.add_data_validation(dvy)
    dvy.add(f"G{first_data}:G{last}")
    dvy.add(f"I{first_data}:J{last}")

    # completeness check, so a half-filled sheet is visible at a glance
    chk = last + 2
    sh.cell(row=chk, column=5, value="Turns with a response recorded:").font = Font(
        name=FONT, bold=True, size=10)
    c = sh.cell(row=chk, column=6,
                value=f'=COUNTA(F{first_data}:F{last})&" of {len(turns)}"')
    c.font = Font(name=FONT, bold=True, size=10)
    sh.cell(row=chk + 1, column=5, value="Turns still blank:").font = Font(name=FONT, size=10)
    sh.cell(row=chk + 1, column=6,
            value=f"=COUNTBLANK(F{first_data}:F{last})").font = Font(name=FONT, size=10)

# ---------------------------------------------------------------- Reference
rf = wb.create_sheet("Reference")
rf.sheet_view.showGridLines = False
ref = [
    ("Scoring categories (filled later, not by you)", "h"),
    ("accurate_complete", "The answer is correct and covers everything the rubric requires."),
    ("substantive_incomplete", "Correct as far as it goes, but misses required detail."),
    ("not_substantive", "Did not answer: said it did not know, or sent you elsewhere."),
    ("incorrect", "Gave wrong information material enough to affect a coverage decision."),
    ("", ""),
    ("Why memory and personalisation must be off", "h"),
    ("", "A product that has learned about you is not answering the question the study asked, "
         "and two sessions with different memory are not comparable with each other."),
    ("", ""),
    ("Why the wording must not change", "h"),
    ("", "The human baseline came from counselors answering these exact questions. Rewording "
         "makes the comparison meaningless, and the rubric may no longer apply."),
    ("", ""),
    ("What 'searched_web' means", "h"),
    ("", "Whether the product looked something up rather than answering from memory. Look for a "
         "'searching the web' indicator, inline source links, or a citation list. In the API "
         "results this is detectable from cost and returned citations, so recording it here "
         "keeps the two comparable."),
]
r = 1
for a, b in ref:
    ca = rf.cell(row=r, column=1, value=a)
    cb = rf.cell(row=r, column=2, value=b)
    if b == "h" or a.endswith(("off", "change", "means", "you)")) and b == "h":
        pass
    if b == "h":
        ca.font = Font(name=FONT, bold=True, size=11, color="1F3864")
        cb.value = None
    else:
        ca.font = Font(name=FONT, bold=True, size=10)
        cb.font = BODY
        cb.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
rf.column_dimensions["A"].width = 26
rf.column_dimensions["B"].width = 95

wb.save(OUT)
print(f"written {OUT}")
print("sheets:", wb.sheetnames)
